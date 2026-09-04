"""Add the VBA project to the built workbook.

1. Builds a minimal workbook with the same sheet names and codenames.
2. Uses LibreOffice (UNO) to insert the VBA modules and export a real vbaProject.bin.
3. Injects that vbaProject.bin into the openpyxl-built workbook and writes the .xlsm.

Usage: python add_vba.py <built.xlsx> <out.xlsm>
Requires a headless soffice listening on port 2002.
"""
import sys, zipfile, re, os, shutil
import openpyxl
import uno
from com.sun.star.beans import PropertyValue

SRC, OUT = sys.argv[1], sys.argv[2]
HERE = os.path.dirname(os.path.abspath(__file__))
SHEET_MODULES = {'Play': 'Sheet1', 'Budget': 'Sheet2', 'Long game': 'Sheet3', 'Giving': 'Sheet4', 'Finish': 'Sheet5'}

# 1. minimal base workbook with the same sheets and codenames
src = openpyxl.load_workbook(SRC)
names = src.sheetnames
codenames = {ws.title: ws.sheet_properties.codeName for ws in src.worksheets}
base = openpyxl.Workbook(); base.remove(base.active)
for n in names:
    ws = base.create_sheet(n)
    ws['A1'] = n
base_path = os.path.join(HERE, 'vba_base.xlsx')
base.save(base_path)

SHEET_CODE = """Option Explicit

Private Sub Worksheet_SelectionChange(ByVal Target As Range)
    Game.HandleClick Me, Target
End Sub
"""
WB_CODE = """Option Explicit

Private Sub Workbook_Open()
    Game.OpenGame
End Sub
"""

def pv(n, v):
    p = PropertyValue(); p.Name = n; p.Value = v; return p

local = uno.getComponentContext()
resolver = local.ServiceManager.createInstanceWithContext("com.sun.star.bridge.UnoUrlResolver", local)
ctx = resolver.resolve("uno:socket,host=127.0.0.1,port=2002;urp;StarOffice.ComponentContext")
desktop = ctx.ServiceManager.createInstanceWithContext("com.sun.star.frame.Desktop", ctx)
doc = desktop.loadComponentFromURL(uno.systemPathToFileUrl(base_path), "_blank", 0, (pv("Hidden", True),))
libs = doc.BasicLibraries
libs.VBACompatibilityMode = True
libs.ProjectName = 'VBAProject'
if not libs.hasByName('VBAProject'):
    libs.createLibrary('VBAProject')
lib = libs.getByName('VBAProject')

def add_module(name, code, mtype, obj=None):
    if lib.hasByName(name):
        lib.removeByName(name)
    lib.insertByName(name, code)
    mi = uno.createUnoStruct('com.sun.star.script.ModuleInfo')
    mi.ModuleType = mtype
    if obj is not None:
        mi.ModuleObject = obj
    lib.insertModuleInfo(name, mi)

add_module('Game', open(os.path.join(HERE, 'Game.bas')).read(), 1)
doc.CodeName = 'ThisWorkbook'
add_module('ThisWorkbook', WB_CODE, 4, doc)
for sheet_name, code in SHEET_MODULES.items():
    sh = doc.Sheets.getByName(sheet_name)
    sh.CodeName = codenames[sheet_name]
    add_module(codenames[sheet_name], SHEET_CODE, 4, sh)
tmp_xlsm = os.path.join(HERE, 'vba_export.xlsm')
if os.path.exists(tmp_xlsm):
    os.remove(tmp_xlsm)
doc.storeToURL(uno.systemPathToFileUrl(tmp_xlsm), (pv("FilterName", "Calc MS Excel 2007 VBA XML"),))
doc.close(True)
vba_bin = zipfile.ZipFile(tmp_xlsm).read('xl/vbaProject.bin')
print('vbaProject.bin bytes:', len(vba_bin))

# 3. inject into the built workbook
zin = zipfile.ZipFile(SRC)
zout = zipfile.ZipFile(OUT, 'w', zipfile.ZIP_DEFLATED)
for item in zin.infolist():
    data = zin.read(item.filename)
    if item.filename == '[Content_Types].xml':
        s = data.decode()
        if 'Extension="bin"' not in s:
            s = s.replace('<Default Extension="rels"', '<Default Extension="bin" ContentType="application/vnd.ms-office.vbaProject"/><Default Extension="rels"', 1)
        s = s.replace('application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml', 'application/vnd.ms-excel.sheet.macroEnabled.main+xml')
        data = s.encode()
    elif item.filename == 'xl/_rels/workbook.xml.rels':
        s = data.decode()
        if 'vbaProject' not in s:
            s = s.replace('</Relationships>', '<Relationship Id="rIdVBA" Type="http://schemas.microsoft.com/office/2006/relationships/vbaProject" Target="vbaProject.bin"/></Relationships>')
        data = s.encode()
    elif item.filename == 'xl/workbook.xml':
        s = data.decode()
        if 'codeName' not in s:
            s = re.sub(r'<workbookPr([^/]*)/>', r'<workbookPr\1 codeName="ThisWorkbook"/>', s, count=1)
        data = s.encode()
    zout.writestr(item, data)
zout.writestr('xl/vbaProject.bin', vba_bin)
zout.close()
# checks
z = zipfile.ZipFile(OUT)
wbxml = z.read('xl/workbook.xml').decode()
print('workbookPr:', re.findall(r'<workbookPr[^>]*>', wbxml))
for n in sorted(x for x in z.namelist() if x.startswith('xl/worksheets/sheet') and x.endswith('.xml'))[:5]:
    print(n, re.findall(r'<sheetPr[^>]*>', z.read(n).decode())[:1])
print('content types has bin:', 'Extension="bin"' in z.read('[Content_Types].xml').decode(), '| macroEnabled:', 'macroEnabled' in z.read('[Content_Types].xml').decode())
print('rels has vba:', 'vbaProject' in z.read('xl/_rels/workbook.xml.rels').decode())
print('saved', OUT)
